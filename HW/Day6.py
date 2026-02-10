"""
HW Day6：RAG AI 客服助手 + DeepEval 評估
==========================================
台灣自來水公司 QA 系統
流程：
  1. qa_data.txt → 滑動視窗切塊 → Qdrant VDB + BM25 索引
  2. Hybrid Search（Dense + BM25） → RRF 融合 → ReRank → LLM 答案
  3. DeepEval 評估 5 項指標
  4. 輸出 day6_HW_questions.csv
"""

import os
import re
import csv
import json
import time
import math
import random
import asyncio
import subprocess
import requests
import numpy as np
import jieba
import openpyxl
from rank_bm25 import BM25Okapi
from langchain_text_splitters import RecursiveCharacterTextSplitter
from qdrant_client import QdrantClient
from qdrant_client.models import Distance, VectorParams, PointStruct

# DeepEval
from openai import OpenAI
from deepeval.models import DeepEvalBaseLLM
from deepeval.metrics import (
    FaithfulnessMetric,
    AnswerRelevancyMetric,
    ContextualRecallMetric,
    ContextualPrecisionMetric,
    ContextualRelevancyMetric,
)
from deepeval.test_case import LLMTestCase


# ============================================================
# 設定
# ============================================================
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))

# API
EMBED_API_URL = "https://ws-04.wade0426.me/embed"
LLM_API_URL = "https://ws-05.huannago.com/v1"
LLM_MODEL = "google/gemma-3-27b-it"

# 切塊參數
CHUNK_SIZE = 500
CHUNK_OVERLAP = 100

# 檢索參數
DENSE_TOP_K = 10
BM25_TOP_K = 10
HYBRID_TOP_K = 10
RERANK_TOP_K = 3
RRF_K = 60

COLLECTION_NAME = "day6_water_qa"

# 抽樣設定（從 30 題中隨機抽 N 題，設為 0 或 30 則全部跑）
SAMPLE_N = 5


# ============================================================
# DeepEval 自訂 LLM
# ============================================================
class CustomLLM(DeepEvalBaseLLM):
    """使用課程提供的 LLM API 作為 DeepEval 評估模型（含重試機制）"""

    def __init__(self, base_url=LLM_API_URL, model_name=LLM_MODEL):
        self.base_url = base_url
        self.model_name = model_name
        self.max_retries = 5          # 最多重試 5 次
        self.base_delay = 10          # 基礎等待 10 秒
        self.call_count = 0           # 追蹤呼叫次數

    def load_model(self):
        return OpenAI(api_key="NoNeed", base_url=self.base_url)

    def generate(self, prompt: str) -> str:
        self.call_count += 1
        client = self.load_model()

        for attempt in range(self.max_retries):
            try:
                response = client.chat.completions.create(
                    model=self.model_name,
                    messages=[{"role": "user", "content": prompt}],
                    temperature=0.1,
                    max_tokens=2048,
                    timeout=180,  # 3 分鐘超時
                )
                content = response.choices[0].message.content

                # 檢查是否回傳了 HTML 錯誤頁面（524 timeout 等）
                if content and "<html" in content.lower()[:100]:
                    raise Exception("API 回傳 HTML 錯誤頁面（可能是 524 Timeout）")

                # 每 5 次呼叫暫停 2 秒，避免過度壓力
                if self.call_count % 5 == 0:
                    time.sleep(2)

                return content or ""

            except Exception as e:
                err_msg = str(e)[:120]
                # 判斷是否為 HTML 錯誤（524 Timeout）
                if "<html" in err_msg.lower() or "524" in err_msg or "timeout" in err_msg.lower():
                    wait = self.base_delay * (2 ** attempt)  # 10, 20, 40, 80, 160 秒
                    print(f"  ⏳ API Timeout（第 {attempt+1}/{self.max_retries} 次），等待 {wait} 秒...")
                    time.sleep(wait)
                else:
                    wait = self.base_delay * (attempt + 1)
                    print(f"  ⚠️ LLM error（第 {attempt+1} 次）: {err_msg}，等待 {wait} 秒...")
                    time.sleep(wait)

        print(f"  ❌ LLM 重試 {self.max_retries} 次後仍失敗")
        return ""

    async def a_generate(self, prompt: str) -> str:
        return self.generate(prompt)

    def get_model_name(self):
        return f"CustomLLM ({self.model_name})"


# ============================================================
# 工具函數
# ============================================================
def get_embedding(texts: list[str]) -> tuple:
    """Embedding API"""
    data = {"texts": texts, "normalize": True, "batch_size": 32}
    for attempt in range(3):
        try:
            resp = requests.post(EMBED_API_URL, json=data, timeout=120)
            if resp.status_code == 200:
                result = resp.json()
                return result["embeddings"], result["dimension"]
            print(f"  ⚠️ Embedding API {resp.status_code}, retry {attempt+1}")
        except Exception as e:
            print(f"  ⚠️ Embedding error: {e}, retry {attempt+1}")
        time.sleep(2)
    return None, None


def call_llm(system_prompt: str, user_prompt: str,
             temperature: float = 0.1, max_tokens: int = 1024) -> str:
    """LLM API"""
    payload = {
        "model": LLM_MODEL,
        "messages": [
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": user_prompt},
        ],
        "temperature": temperature,
        "max_tokens": max_tokens,
    }
    for attempt in range(3):
        try:
            resp = requests.post(f"{LLM_API_URL}/chat/completions",
                                 json=payload, timeout=120)
            if resp.status_code == 200:
                return resp.json()["choices"][0]["message"]["content"].strip()
            print(f"  ⚠️ LLM {resp.status_code}, retry {attempt+1}")
        except Exception as e:
            print(f"  ⚠️ LLM error: {e}, retry {attempt+1}")
        time.sleep(2)
    return ""


# ============================================================
# 文本切塊
# ============================================================
def chunk_qa_data(text: str) -> list[dict]:
    """將 QA 資料切塊"""
    splitter = RecursiveCharacterTextSplitter(
        separators=["\n\n\n", "\n\n", "\n", "。", "！", "？", "；", "，", ""],
        chunk_size=CHUNK_SIZE,
        chunk_overlap=CHUNK_OVERLAP,
        length_function=len,
    )
    chunks = splitter.split_text(text)
    result = []
    for c in chunks:
        result.append({"text": c, "source": "qa_data.txt"})
    return result


# ============================================================
# BM25 索引
# ============================================================
def tokenize_chinese(text: str) -> list[str]:
    """中文分詞"""
    words = jieba.lcut(text)
    stop_words = {"的", "了", "在", "是", "我", "有", "和", "就",
                  "不", "人", "都", "一", "一個", "上", "也", "很",
                  "到", "說", "要", "去", "你", "會", "著", "沒有",
                  "看", "好", "自己", "這", "他", "她", "它", "們",
                  "那", "被", "從", "對", "為", "與", "等", "但",
                  "而", "及", "或", "之", "其", "中", "所", "以",
                  "可", "能", "將", "還", "因", "此", "則", "如",
                  "於", "個", "每", "又", "把", "讓", "用", "做",
                  "嗎", "呢", "啊", "吧", "喔", "呀", "耶"}
    return [w for w in words if len(w) > 1 and w not in stop_words]


class BM25Index:
    def __init__(self, chunks):
        self.chunks = chunks
        self.tokenized = [tokenize_chinese(c["text"]) for c in chunks]
        self.bm25 = BM25Okapi(self.tokenized)

    def search(self, query, top_k=BM25_TOP_K):
        tokens = tokenize_chinese(query)
        scores = self.bm25.get_scores(tokens)
        top_idx = np.argsort(scores)[::-1][:top_k]
        results = []
        for idx in top_idx:
            if scores[idx] > 0:
                results.append({
                    "index": int(idx),
                    "text": self.chunks[idx]["text"],
                    "source": self.chunks[idx]["source"],
                    "bm25_score": float(scores[idx]),
                })
        return results


# ============================================================
# Hybrid Search + RRF
# ============================================================
def rrf_fusion(dense_results, bm25_results, k=RRF_K, top_k=HYBRID_TOP_K):
    rrf_scores = {}
    doc_info = {}
    for rank, hit in enumerate(dense_results):
        idx = hit["index"]
        rrf_scores[idx] = rrf_scores.get(idx, 0) + 1.0 / (k + rank + 1)
        doc_info[idx] = {"text": hit["text"], "source": hit["source"]}
    for rank, hit in enumerate(bm25_results):
        idx = hit["index"]
        rrf_scores[idx] = rrf_scores.get(idx, 0) + 1.0 / (k + rank + 1)
        doc_info[idx] = {"text": hit["text"], "source": hit["source"]}
    sorted_docs = sorted(rrf_scores.items(), key=lambda x: x[1], reverse=True)
    return [{"index": idx, "text": doc_info[idx]["text"],
             "source": doc_info[idx]["source"], "rrf_score": sc}
            for idx, sc in sorted_docs[:top_k]]


# ============================================================
# ReRank
# ============================================================
RERANK_SYSTEM = """你是一個文件相關性評估專家。請判斷「文件段落」與「查詢問題」的相關程度。
評分標準（0-10 分）：
- 9-10：文件直接且完整地回答了問題
- 7-8：文件包含大部分關鍵資訊
- 5-6：文件部分相關
- 3-4：略有相關
- 0-2：完全無關
請只輸出一個整數（0-10），不要有任何其他文字。"""


def rerank(query, candidates, top_k=RERANK_TOP_K):
    scored = []
    for cand in candidates:
        prompt = f"【問題】{query}\n\n【文件】{cand['text']}"
        resp = call_llm(RERANK_SYSTEM, prompt, temperature=0.0, max_tokens=16)
        try:
            nums = re.findall(r'\d+', resp)
            score = min(max(float(nums[0]), 0), 10) if nums else 5.0
        except:
            score = 5.0
        scored.append({**cand, "rerank_score": score})
        time.sleep(0.1)
    scored.sort(key=lambda x: x["rerank_score"], reverse=True)
    return scored[:top_k]


# ============================================================
# Query ReWrite（口語化 → 搜尋語句）
# ============================================================
REWRITE_SYSTEM = """你是一個 RAG 查詢重寫專家。請將使用者的口語化問題重寫為適合向量資料庫搜尋的精確查詢語句。

規則：
1. 將口語表達轉為正式用語（例如「那個紙張的單子」→「紙本帳單」）
2. 補充關鍵詞（例如「白白的」→「白濁 空氣 氣泡」）
3. 保留原意，不要回答問題
4. 使用繁體中文
5. 只輸出重寫後的語句，不要解釋"""


def rewrite_query(question: str) -> str:
    """口語化問題 → 搜尋語句"""
    rewritten = call_llm(REWRITE_SYSTEM, question, temperature=0.1, max_tokens=128)
    return rewritten.strip() if rewritten else question


# ============================================================
# LLM 答案生成
# ============================================================
ANSWER_SYSTEM = """你是台灣自來水公司的 AI 客服助手。請根據「參考資料」精準回答用戶的問題。

規則：
1. 只根據參考資料中的內容回答，不要編造
2. 回答要親切、專業、完整，適合一般民眾理解
3. 包含所有相關的關鍵資訊（金額、期限、流程等）
4. 使用繁體中文
5. 不要加「根據參考資料」等前綴，直接回答"""


def generate_answer(question, chunks):
    context = ""
    for i, c in enumerate(chunks):
        context += f"【資料 {i+1}】\n{c['text']}\n\n"
    prompt = f"【參考資料】\n{context}\n【用戶問題】\n{question}\n\n請回答："
    return call_llm(ANSWER_SYSTEM, prompt)


# ============================================================
# 完整 RAG Pipeline
# ============================================================
def rag_pipeline(query, qdrant_client, bm25_index, chunks,
                 use_rewrite=True) -> tuple[str, list[str]]:
    """
    完整 RAG 流程
    回傳 (answer, retrieval_context_list)
    """
    # Query ReWrite
    search_query = rewrite_query(query) if use_rewrite else query

    # Dense Search
    emb, _ = get_embedding([search_query])
    if emb is None:
        return "", []

    dense_raw = qdrant_client.query_points(
        collection_name=COLLECTION_NAME, query=emb[0], limit=DENSE_TOP_K
    )
    dense_results = [{"index": p.id, "text": p.payload["text"],
                      "source": p.payload["source"], "dense_score": p.score}
                     for p in dense_raw.points]

    # BM25 Search
    bm25_results = bm25_index.search(search_query, BM25_TOP_K)

    # RRF Fusion
    hybrid = rrf_fusion(dense_results, bm25_results, top_k=HYBRID_TOP_K)

    # ReRank
    reranked = rerank(query, hybrid, RERANK_TOP_K)

    # 收集 context（DeepEval 需要）
    retrieval_context = [c["text"] for c in reranked]

    # LLM Answer
    answer = generate_answer(query, reranked)

    return answer, retrieval_context


# ============================================================
# DeepEval 評估
# ============================================================
def evaluate_with_deepeval(question, answer, expected_answer,
                           retrieval_context, eval_llm):
    """使用 DeepEval 評估 5 項指標"""

    test_case = LLMTestCase(
        input=question,
        actual_output=answer,
        expected_output=expected_answer,
        retrieval_context=retrieval_context,
    )

    metrics = {
        "Faithfulness": FaithfulnessMetric(model=eval_llm, threshold=0.5),
        "Answer_Relevancy": AnswerRelevancyMetric(model=eval_llm, threshold=0.5),
        "Contextual_Recall": ContextualRecallMetric(model=eval_llm, threshold=0.5),
        "Contextual_Precision": ContextualPrecisionMetric(model=eval_llm, threshold=0.5),
        "Contextual_Relevancy": ContextualRelevancyMetric(model=eval_llm, threshold=0.5),
    }

    scores = {}
    for name, metric in metrics.items():
        try:
            metric.measure(test_case)
            scores[name] = round(metric.score, 4)
        except Exception as e:
            err_msg = str(e)[:120]
            print(f"    ⚠️ {name} 評估失敗: {err_msg}")
            # 如果是 timeout/JSON 錯誤，等待後重試一次
            if "timeout" in err_msg.lower() or "JSON" in err_msg or "524" in err_msg:
                print(f"    🔄 等待 15 秒後重試 {name}...")
                time.sleep(15)
                try:
                    metric.measure(test_case)
                    scores[name] = round(metric.score, 4)
                    print(f"    ✅ {name} 重試成功: {scores[name]}")
                except Exception as e2:
                    print(f"    ❌ {name} 重試仍失敗: {str(e2)[:80]}")
                    scores[name] = 0.0
            else:
                scores[name] = 0.0
        time.sleep(2)  # 每個指標間隔 2 秒

    return scores


# ============================================================
# 主程式
# ============================================================
def main():
    print("=" * 65)
    print("HW Day6：RAG AI 客服助手 + DeepEval 評估")
    print("台灣自來水公司 QA 系統")
    print("=" * 65)

    # ── 1. 讀取資料 ──
    print("\n📂 步驟 1：讀取資料")
    print("-" * 40)

    # 讀取 qa_data（支援 .txt 和 .docx）
    qa_txt_path = os.path.join(SCRIPT_DIR, "qa_data.txt")
    qa_docx_path = os.path.join(SCRIPT_DIR, "qa_data.docx")
    if os.path.exists(qa_txt_path):
        with open(qa_txt_path, "r", encoding="utf-8") as f:
            qa_text = f.read()
    elif os.path.exists(qa_docx_path):
        # 用 pandoc 將 docx 轉為純文字
        result = subprocess.run(
            ["pandoc", qa_docx_path, "-t", "plain", "--wrap=none"],
            capture_output=True, text=True
        )
        qa_text = result.stdout
        # 也存一份 txt 供後續使用
        with open(qa_txt_path, "w", encoding="utf-8") as f:
            f.write(qa_text)
    else:
        print("❌ 找不到 qa_data.txt 或 qa_data.docx")
        return
    print(f"  ✅ qa_data：{len(qa_text)} 字元")

    # 讀取 questions（支援 .csv 和 .xlsx）
    q_csv_path = os.path.join(SCRIPT_DIR, "questions.csv")
    q_xlsx_path = os.path.join(SCRIPT_DIR, "day6_HW_questions.csv.xlsx")
    if os.path.exists(q_csv_path):
        with open(q_csv_path, "r", encoding="utf-8-sig") as f:
            questions = list(csv.DictReader(f))
    elif os.path.exists(q_xlsx_path):
        questions = []
        wb = openpyxl.load_workbook(q_xlsx_path)
        ws = wb.active
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        for row in ws.iter_rows(min_row=2, values_only=True):
            d = {}
            for h, v in zip(headers, row):
                if isinstance(v, float) and v == int(v):
                    d[h] = int(v)
                elif v is None:
                    d[h] = ""
                else:
                    d[h] = v
            questions.append(d)
    else:
        print("❌ 找不到 questions.csv 或 day6_HW_questions.csv.xlsx")
        return
    print(f"  ✅ questions：{len(questions)} 題")

    # 讀取參考答案（支援 .csv 和 .xlsx）
    qa_ans_csv = os.path.join(SCRIPT_DIR, "questions_answer.csv")
    qa_ans_xlsx = os.path.join(SCRIPT_DIR, "questions_answer.csv.xlsx")
    ref_answers = {}
    if os.path.exists(qa_ans_csv):
        with open(qa_ans_csv, "r", encoding="utf-8-sig") as f:
            for row in csv.DictReader(f):
                ref_answers[int(float(row["q_id"]))] = row["answer"]
    elif os.path.exists(qa_ans_xlsx):
        wb = openpyxl.load_workbook(qa_ans_xlsx)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            q_id = int(float(row[0])) if row[0] else 0
            answer = row[2] if row[2] else ""
            ref_answers[q_id] = answer
    else:
        print("⚠️ 找不到參考答案檔案，DeepEval 部分指標可能受影響")
    print(f"  ✅ 參考答案：{len(ref_answers)} 題")

    # 隨機抽樣 N 題
    if 0 < SAMPLE_N < len(questions):
        random.seed(42)  # 固定 seed 讓結果可重現
        questions = random.sample(questions, SAMPLE_N)
        sampled_ids = [int(float(q["q_id"])) for q in questions]
        print(f"\n  🎲 隨機抽樣 {SAMPLE_N} 題：Q{sampled_ids}")

    # ── 2. 切塊 ──
    print(f"\n📦 步驟 2：滑動視窗切塊（size={CHUNK_SIZE}, overlap={CHUNK_OVERLAP}）")
    print("-" * 40)
    chunks = chunk_qa_data(qa_text)
    print(f"  ✅ 共 {len(chunks)} 個切塊")

    # ── 3. Qdrant + BM25 ──
    print(f"\n🔗 步驟 3：建立向量資料庫 + BM25 索引")
    print("-" * 40)

    _, dim = get_embedding(["測試"])
    print(f"  ✅ Embedding 維度：{dim}")

    client = QdrantClient(url="http://localhost:6333")

    existing = [c.name for c in client.get_collections().collections]
    if COLLECTION_NAME in existing:
        client.delete_collection(COLLECTION_NAME)
    client.create_collection(
        collection_name=COLLECTION_NAME,
        vectors_config=VectorParams(size=dim, distance=Distance.COSINE),
    )

    all_points = []
    batch_size = 50
    for i in range(0, len(chunks), batch_size):
        batch = chunks[i:i + batch_size]
        embs, _ = get_embedding([c["text"] for c in batch])
        if embs is None:
            continue
        for j, (chunk, emb) in enumerate(zip(batch, embs)):
            all_points.append(PointStruct(
                id=i + j, vector=emb,
                payload={"text": chunk["text"], "source": chunk["source"]},
            ))
        time.sleep(0.3)

    client.upsert(collection_name=COLLECTION_NAME, points=all_points)
    print(f"  ✅ Qdrant：{len(all_points)} 個向量")

    bm25_index = BM25Index(chunks)
    print(f"  ✅ BM25 索引建立完成")

    # ── 4. RAG 回答 30 題 ──
    print(f"\n🔍 步驟 4：RAG 回答 {len(questions)} 題")
    print(f"  流程：Query ReWrite → Hybrid Search → ReRank → LLM Answer")
    print("-" * 40)

    # 嘗試載入已完成的 RAG 結果（避免重複呼叫）
    rag_checkpoint_path = os.path.join(SCRIPT_DIR, "rag_checkpoint.json")
    results = []
    loaded_rag = {}
    if os.path.exists(rag_checkpoint_path):
        with open(rag_checkpoint_path, "r", encoding="utf-8") as f:
            loaded_rag = {str(r["q_id"]): r for r in json.load(f)}
        print(f"  📌 找到 RAG 進度檔，已完成 {len(loaded_rag)} 題")

    for q in questions:
        q_id = int(float(q["q_id"]))
        q_text = q["questions"]
        q_key = str(q_id)

        # 如果已有 RAG 結果，跳過
        if q_key in loaded_rag:
            r = loaded_rag[q_key]
            print(f"\n  ⏭️ Q{q_id}: 使用已有結果")
            results.append(r)
            continue

        print(f"\n  Q{q_id}: {q_text[:50]}...")

        answer, context = rag_pipeline(q_text, client, bm25_index, chunks)
        print(f"    ✅ 答案：{answer[:60]}...")

        results.append({
            "q_id": q_id,
            "questions": q_text,
            "answer": answer,
            "context": context,
            "expected": ref_answers.get(q_id, ""),
        })
        time.sleep(0.5)

    # 儲存 RAG 結果
    with open(rag_checkpoint_path, "w", encoding="utf-8") as f:
        json.dump(results, f, ensure_ascii=False, indent=2)
    print(f"\n  💾 RAG 結果已儲存至 rag_checkpoint.json")

    # ── 5. DeepEval 評估 ──
    print(f"\n📊 步驟 5：DeepEval 評估 5 項指標")
    print("-" * 40)

    eval_llm = CustomLLM()
    print(f"  ✅ 評估模型：{eval_llm.get_model_name()}")

    # 載入已有的評估進度（斷點續跑）
    checkpoint_path = os.path.join(SCRIPT_DIR, "eval_checkpoint.json")
    existing_scores = {}
    if os.path.exists(checkpoint_path):
        with open(checkpoint_path, "r", encoding="utf-8") as f:
            existing_scores = json.load(f)
        print(f"  📌 找到評估進度檔，已完成 {len(existing_scores)} 題，從斷點繼續...")

    for i, r in enumerate(results):
        q_id = r["q_id"]
        q_key = str(q_id)

        # 如果已有評估結果，跳過
        if q_key in existing_scores:
            r["scores"] = existing_scores[q_key]
            print(f"\n  ⏭️ Q{q_id} 已有評估結果，跳過")
            for name, val in r["scores"].items():
                print(f"    {name}: {val}")
            continue

        print(f"\n  📊 評估 Q{q_id}（{i+1}/{len(results)}）...")

        # 每題評估前暫停，避免 API 過載
        if i > 0:
            delay = 5  # 每題間隔 5 秒
            print(f"  ⏳ 等待 {delay} 秒避免 API 過載...")
            time.sleep(delay)

        scores = evaluate_with_deepeval(
            question=r["questions"],
            answer=r["answer"],
            expected_answer=r["expected"],
            retrieval_context=r["context"],
            eval_llm=eval_llm,
        )

        r["scores"] = scores
        for name, val in scores.items():
            print(f"    {name}: {val}")

        # 即時儲存進度（斷點）
        existing_scores[q_key] = scores
        with open(checkpoint_path, "w", encoding="utf-8") as f:
            json.dump(existing_scores, f, ensure_ascii=False, indent=2)

    # 評估完成，刪除 checkpoint
    if os.path.exists(checkpoint_path):
        os.remove(checkpoint_path)
    if os.path.exists(rag_checkpoint_path):
        os.remove(rag_checkpoint_path)
    print(f"\n  ✅ 全部評估完成，已清理 checkpoint 檔案")

    # ── 6. 輸出 CSV ──
    print(f"\n💾 步驟 6：輸出 day6_HW_questions.csv")
    print("-" * 40)

    csv_path = os.path.join(SCRIPT_DIR, "day6_HW_questions.csv")
    with open(csv_path, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.writer(f)
        writer.writerow([
            "q_id", "questions", "answer",
            "Faithfulness", "Answer_Relevancy",
            "Contextual_Recall", "Contextual_Precision", "Contextual_Relevancy"
        ])
        for r in results:
            s = r.get("scores", {})
            writer.writerow([
                r["q_id"], r["questions"], r["answer"],
                s.get("Faithfulness", 0),
                s.get("Answer_Relevancy", 0),
                s.get("Contextual_Recall", 0),
                s.get("Contextual_Precision", 0),
                s.get("Contextual_Relevancy", 0),
            ])
    print(f"  ✅ {csv_path}")

    # ── 7. 摘要 ──
    print(f"""
{'=' * 65}
✅ HW Day6 完成！
{'=' * 65}

📋 系統架構：
  資料來源：qa_data.txt（{len(qa_text)} 字元 → {len(chunks)} 塊）
  切塊方法：滑動視窗（size={CHUNK_SIZE}, overlap={CHUNK_OVERLAP}）
  Dense Search：Qdrant VDB（{dim} 維）
  Sparse Search：BM25（jieba 中文分詞）
  Hybrid Fusion：RRF（k={RRF_K}）
  ReRank：LLM 相關性評分
  Answer Gen：{LLM_MODEL}
  Evaluation：DeepEval（5 metrics）

📊 DeepEval 平均分數：""")

    metric_names = ["Faithfulness", "Answer_Relevancy",
                    "Contextual_Recall", "Contextual_Precision",
                    "Contextual_Relevancy"]
    for m in metric_names:
        vals = [r["scores"].get(m, 0) for r in results if "scores" in r]
        avg = sum(vals) / len(vals) if vals else 0
        print(f"  {m}: {avg:.4f}")


if __name__ == "__main__":
    main()